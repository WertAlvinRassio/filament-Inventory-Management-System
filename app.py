#!/usr/bin/env python3
"""FilaWizard Filament Inventory (Rebuilt Backend)

- Uses adafruit_tca9548a for PCA9548 I2C mux management
- Uses cedargrove_nau7802 driver for NAU7802 24-bit ADC
- Uses adafruit_hdc302x for HDC302x temperature/humidity sensor
- Uses adafruit_pn532 for PN532 NFC via NAU7802 second STEMMA QT port
- Auto-detects active mux channels and disables unused connections
- Channel 7: HDC302x environmental sensor
- Channels 0-6: NAU7802 load cell + PN532 NFC reader
- Keeps existing UI + NFC format + calibration + Excel export + history
- Prevents silent 500s: API endpoints return JSON with error details

NFC payload format:
  Brand|Color|Type|Tare
"""

import io
import json
import logging
import os
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Optional, Tuple, Dict, Any, List

from flask import Flask, jsonify, render_template, request, send_file

# Hardware libs (Pi)
import board
import busio
import adafruit_tca9548a
import adafruit_hdc302x
from cedargrove_nau7802 import NAU7802
from adafruit_pn532.i2c import PN532_I2C

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger("filawizard")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    stream=sys.stderr,
)

# ---------------------------
# Configuration
# ---------------------------

APP_TITLE = "FilaWizard"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "inventory_config.json")
HISTORY_FILE = os.path.join(BASE_DIR, "history.jsonl")

PCA9548_ADDRESSES = [0x70, 0x71, 0x72, 0x73, 0x74, 0x75, 0x76, 0x77]

# Environmental sensor wiring
HDC302X_ADDR = 0x44
HDC302X_MUX_INDEX = 0
HDC302X_CHANNEL = 7

# Device addresses
NAU7802_ADDR = 0x2A
PN532_ADDR = 0x24

# Business logic
LOW_FILAMENT_THRESHOLD_G = 250
NEW_ROLL_DEFAULT_G = 1000

# Loop
SCAN_INTERVAL_SEC = 2.0

# Stable muxed NAU7802 read tuning
NAU_WARMUP_READS = 30
NAU_FILTER_SAMPLES = 40
NAU_SAMPLE_DELAY_SEC = 0.03

# Default slots: 48 (6 muxes x 8 channels). If fewer muxes are connected, slots reduce automatically.
DEFAULT_SLOT_COUNT = int(os.environ.get("SLOT_COUNT", "48"))

# ---------------------------
# Helpers
# ---------------------------

def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def safe_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def clamp_text(s: str, max_len: int = 32) -> str:
    return (s or "").strip()[:max_len]

def calibration_status(last_calibrated: Optional[str]) -> str:
    if not last_calibrated:
        return "Not calibrated"
    try:
        from datetime import timezone, timedelta
        s = last_calibrated.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        if datetime.now(timezone.utc) - dt > timedelta(days=365):
            return "Not calibrated"
        return "Calibrated"
    except Exception:
        return "Not calibrated"

def append_history(event: Dict[str, Any]) -> None:
    try:
        with open(HISTORY_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(event, ensure_ascii=False) + "\n")
    except Exception:
        pass

# ---------------------------
# Load cell (NAU7802 via cedargrove driver)
# ---------------------------

class NAU7802LoadCell:
    """Wrapper around cedargrove_nau7802 NAU7802 driver for load cell reading
    through a TCA9548A mux channel.  The TCA9548A driver handles mux channel
    switching transparently when devices are created with tca[channel]."""

    def __init__(self, tca_channel, address=NAU7802_ADDR):
        self.tca_channel = tca_channel
        self.address = address
        self.available = False
        self.initialized = False
        self.last_error = None
        self._nau = None
        self._init_device()

    def _init_device(self):
        """Initialize the NAU7802 using the cedargrove driver.
        The constructor handles: reset, enable, LDO 3V0, gain 128, 10 SPS."""
        try:
            self._nau = NAU7802(self.tca_channel, address=self.address, active_channels=1)
            # Perform internal offset calibration
            if not self._nau.calibrate("INTERNAL"):
                self.last_error = "NAU7802 internal calibration failed"
                return

            self.available = True
            self.initialized = True

            # Discard warmup reads to let the ADC stabilize
            for _ in range(10):
                if self._nau.available():
                    self._nau.read()
                time.sleep(0.01)

        except Exception as e:
            self.last_error = str(e)
            self.available = False
            self.initialized = False
            self._nau = None

    def data_ready(self) -> bool:
        """Check if ADC conversion data is ready."""
        if not self._nau:
            return False
        try:
            return self._nau.available()
        except Exception:
            return False

    def read_raw_once(self) -> Optional[float]:
        """Read a single ADC value.  Returns signed value or None."""
        if not (self.available and self.initialized and self._nau):
            return None
        try:
            timeout = 100
            while not self._nau.available() and timeout > 0:
                time.sleep(0.02)
                timeout -= 1
            if timeout <= 0:
                return None
            return self._nau.read()
        except Exception as e:
            self.last_error = str(e)
            return None

    def read_filtered_raw(self) -> Optional[float]:
        """Read ADC with warmup and median filtering for stable results."""
        if not (self.available and self.initialized):
            return None

        # Warmup reads (discarded)
        for _ in range(NAU_WARMUP_READS):
            self.read_raw_once()
            time.sleep(NAU_SAMPLE_DELAY_SEC)

        # Collect filtered samples
        vals = []
        for _ in range(NAU_FILTER_SAMPLES):
            v = self.read_raw_once()
            if v is not None:
                vals.append(v)
            time.sleep(NAU_SAMPLE_DELAY_SEC)

        if not vals:
            return None

        vals.sort()
        return vals[len(vals) // 2]

# ---------------------------
# Slot + Config
# ---------------------------

@dataclass
class SlotConfig:
    # counts-per-gram (raw delta / grams)
    calibration_factor: float = 1.0
    # raw reading at zero load
    zero_offset_raw: float = 0.0
    last_calibrated: Optional[str] = None

class FilamentSlot:
    """Represents a single filament slot with NAU7802 load cell and PN532 NFC,
    accessed through a TCA9548A mux channel.  The PN532 is connected via the
    NAU7802's second STEMMA QT passthrough port on the same mux channel."""

    def __init__(self, slot_id: int, tca_channel, cfg: SlotConfig):
        self.slot_id = slot_id
        self.tca_channel = tca_channel
        self.cfg = cfg

        # Initialize NAU7802 load cell via cedargrove driver
        self.nau7802 = NAU7802LoadCell(tca_channel=tca_channel, address=NAU7802_ADDR)
        self.nfc: Optional[PN532_I2C] = None

        self.has_scale = bool(self.nau7802.available)
        self.nfc_available = False
        self.has_nfc = False
        self.hardware_present = bool(self.has_scale)

        # Hotplug tracking (RAM-only)
        self._was_present = self.hardware_present
        self._disconnect_seen = False

        # NFC / material state
        self.uid: Optional[str] = None
        self.filament_brand = "Unknown"
        self.filament_color = "N/A"
        self.filament_type = "N/A"
        self.tag_tare: Optional[float] = None
        self.nfc_needs_programming = False

        self.gross_weight = 0.0
        self.weight = 0.0
        self.is_active = False
        self.last_error: Optional[str] = None

        # Initialize PN532 NFC on the same mux channel (via NAU7802 second STEMMA QT)
        self._init_nfc()

    @property
    def calibration_factor(self) -> float:
        return float(self.cfg.calibration_factor or 1.0)

    def _init_nfc(self) -> None:
        """Initialize PN532 NFC reader on the same mux channel as the NAU7802.
        The PN532 is connected via the NAU7802's second STEMMA QT passthrough."""
        try:
            self.nfc = PN532_I2C(self.tca_channel, debug=False, address=PN532_ADDR)
            self.nfc.SAM_configuration()
            self.nfc_available = True
            self.has_nfc = True
        except Exception as e:
            self.nfc = None
            self.nfc_available = False
            self.has_nfc = False
            self.last_error = f"NFC init: {e}"
        finally:
            self.hardware_present = bool(self.has_scale or self.has_nfc)
            self._was_present = self.hardware_present

    def _read_ntag_text(self) -> Optional[str]:
        """Read text payload from NTAG2xx pages 4-19 (64 bytes).
        Each ntag2xx_read_block returns 4 bytes (one NTAG2xx page)."""
        if not self.nfc_available or not self.nfc:
            return None
        try:
            raw = bytearray()
            for block in range(4, 20):
                data = self.nfc.ntag2xx_read_block(block)
                if not data:
                    break
                raw.extend(bytes(data))
            return raw.decode("utf-8", errors="ignore").strip("\x00").strip()
        except Exception:
            return None

    def _write_ntag_text(self, text: str) -> bool:
        """Write text payload to NTAG2xx pages 4-19 (64 bytes max).
        Each NTAG2xx page is 4 bytes; ntag2xx_write_block expects exactly 4 bytes."""
        if not self.nfc_available or not self.nfc:
            return False
        payload = (text or "").encode("utf-8")[:64]
        payload = payload + b"\x00" * (64 - len(payload))
        try:
            for i in range(16):
                chunk = payload[i * 4:(i + 1) * 4]
                ok = self.nfc.ntag2xx_write_block(4 + i, bytearray(chunk))
                if not ok:
                    return False
            return True
        except Exception:
            return False

    def refresh_hardware(self) -> None:
        """Re-probe hardware and handle hotplug events."""
        prev = self._was_present

        # Re-init scale if it went away
        if not self.nau7802.available:
            self.nau7802 = NAU7802LoadCell(tca_channel=self.tca_channel, address=NAU7802_ADDR)
        self.has_scale = bool(self.nau7802.available)

        # Re-init NFC if it went away
        if not self.nfc_available:
            self._init_nfc()
        self.has_nfc = bool(self.nfc_available)

        self.hardware_present = bool(self.has_scale or self.has_nfc)

        if prev and not self.hardware_present:
            self._disconnect_seen = True
        if self._disconnect_seen and (not prev) and self.hardware_present:
            self.cfg.last_calibrated = None
            self._disconnect_seen = False

        self._was_present = self.hardware_present

    def read_nfc(self) -> bool:
        """Read NFC tag UID and payload.  The TCA9548A driver handles mux
        channel selection transparently."""
        if not self.nfc_available or not self.nfc:
            self.uid = None
            self.nfc_needs_programming = False
            self.has_nfc = False
            self.hardware_present = bool(self.has_scale or self.has_nfc)
            return False

        try:
            uid_bytes = self.nfc.read_passive_target(timeout=0.4)
            if not uid_bytes:
                self.uid = None
                self.nfc_needs_programming = False
                if abs(self.weight) < 10:
                    self.is_active = False
                return False

            self.uid = "".join(f"{b:02X}" for b in uid_bytes)
            self.is_active = True

            txt = self._read_ntag_text()
            if not txt:
                self.nfc_needs_programming = True
                return True

            parts = [p.strip() for p in txt.split("|")]
            parsed_any = False
            if len(parts) >= 3 and parts[0] and parts[1] and parts[2]:
                self.filament_brand, self.filament_color, self.filament_type = parts[0], parts[1], parts[2]
                parsed_any = True

            tare = None
            if len(parts) >= 4 and parts[3] != "":
                tare = safe_float(parts[3])
            self.tag_tare = tare
            self.nfc_needs_programming = not parsed_any
            return True
        except Exception as e:
            self.last_error = f"NFC read: {e}"
            return False

    def read_weight(self):
        """Read filtered weight from the NAU7802 load cell."""
        if not self.nau7802.available:
            self.weight = 0.0
            self.gross_weight = 0.0
            return 0.0

        raw = self.nau7802.read_filtered_raw()

        if raw is None:
            self.weight = 0.0
            self.gross_weight = 0.0
            return 0.0

        counts_per_g = float(self.calibration_factor or 0.0)
        zero = float(self.cfg.zero_offset_raw or 0.0)

        gross = 0.0

        if counts_per_g != 0:
            gross = (raw - zero) / counts_per_g

        if gross < 0 and gross > -50:
            gross = 0.0

        self.gross_weight = float(gross)

        tare_g = float(self.tag_tare) if self.tag_tare else 0.0

        self.weight = float(self.gross_weight - tare_g)

        if abs(self.gross_weight) > 10 or self.uid:
            self.is_active = True

        return self.weight

    def write_tag(self, brand: str, color: str, filament_type: str, tare: Optional[float]) -> bool:
        if not self.uid:
            return False

        brand = clamp_text(brand or self.filament_brand or "Unknown")
        color = clamp_text(color or self.filament_color or "N/A")
        filament_type = clamp_text(filament_type or self.filament_type or "N/A")

        if tare is None:
            tare = self.tag_tare

        tare_str = ""
        if tare is not None:
            try:
                tare_str = str(round(float(tare), 2))
            except Exception:
                tare_str = ""

        payload = f"{brand}|{color}|{filament_type}|{tare_str}"
        ok = self._write_ntag_text(payload)
        if ok:
            self.filament_brand = brand
            self.filament_color = color
            self.filament_type = filament_type
            self.tag_tare = safe_float(tare_str) if tare_str != "" else None
            self.nfc_needs_programming = False
            self.is_active = True
        return ok

    def set_new_roll(self, full_roll_g: float, manual_tare: Optional[float] = None) -> bool:
        if manual_tare is not None:
            tare = max(0.0, float(manual_tare))
        else:
            tare = max(0.0, float(self.gross_weight) - float(full_roll_g))
        return self.write_tag(self.filament_brand, self.filament_color, self.filament_type, tare)

    def get_data(self) -> Dict[str, Any]:
        cal_stat = calibration_status(self.cfg.last_calibrated)
        return {
            "slot_id": self.slot_id,
            "hardware_present": bool(self.hardware_present),
            "has_scale": bool(self.has_scale),
            "has_nfc": bool(self.has_nfc),
            "uid": self.uid,
            "paired_uid": None,
            "requires_new_roll_check": bool(self.gross_weight >= 10 and (self.uid is None or self.nfc_needs_programming)),
            "nfc_needs_programming": bool(self.nfc_needs_programming),
            "brand": self.filament_brand,
            "color": self.filament_color,
            "type": self.filament_type,
            "tare": self.tag_tare,
            "gross_weight": round(float(self.gross_weight or 0.0), 1),
            "weight": round(float(self.weight or 0.0), 1),
            "is_active": bool(self.is_active),
            "calibration_factor": float(self.calibration_factor),
            "last_calibrated": self.cfg.last_calibrated,
            "calibration_status": cal_stat,
            "calibration_required": (cal_stat != "Calibrated"),
            "last_error": self.last_error,
        }

# ---------------------------
# Inventory Manager
# ---------------------------

class InventoryManager:
    """Manages all filament slots across one or more PCA9548 muxes.

    On startup:
      1. Probes PCA9548_ADDRESSES to find connected muxes (via adafruit_tca9548a).
      2. Scans every channel on every mux for known device addresses.
      3. Builds FilamentSlot objects only for channels with a NAU7802.
      4. Creates an HDC302x sensor instance for the designated env channel.
      5. Disables (deselects) all channels that have no active devices.
    """

    def __init__(self, i2c: busio.I2C):
        self.lock = threading.Lock()
        self.i2c = i2c

        # Discover connected PCA9548 muxes
        self.muxes: List[adafruit_tca9548a.TCA9548A] = []
        for addr in PCA9548_ADDRESSES:
            try:
                tca = adafruit_tca9548a.TCA9548A(i2c, address=addr)
                self.muxes.append(tca)
                log.info("Found PCA9548 mux at 0x%02X", addr)
            except Exception as e:
                log.debug("No mux at 0x%02X: %s", addr, e)

        # Active channel map: (mux_index, channel) -> list of detected device names
        self.active_channels: Dict[Tuple[int, int], List[str]] = {}
        self._scan_all_channels()

        # Environmental sensor
        self.sensor: Optional[adafruit_hdc302x.HDC302x] = None
        self.temperature_c: Optional[float] = None
        self.humidity: Optional[float] = None
        self._init_env_sensor()

        # Config + slots
        self.config = self._load_config()
        self.slots: List[FilamentSlot] = []
        self._build_slots()

        # Disable unused channels
        self._disable_unused_channels()

        self.last_error: Optional[str] = None

    # -- Channel detection --

    def _probe_address(self, tca_channel, address: int) -> bool:
        """Attempt a 0-byte write to *address* on *tca_channel* to detect a device.
        TCA9548A_Channel implements the I2C interface (try_lock/unlock/writeto)
        rather than context-manager protocol."""
        try:
            deadline = time.monotonic() + 2.0
            while not tca_channel.try_lock():
                if time.monotonic() > deadline:
                    log.warning("_probe_address: timeout acquiring lock for 0x%02X", address)
                    return False
                time.sleep(0.01)
            try:
                tca_channel.writeto(address, b"")
                return True
            except OSError:
                return False
            finally:
                tca_channel.unlock()
        except Exception as e:
            log.debug("_probe_address 0x%02X exception: %s", address, e)
            return False

    def _scan_all_channels(self) -> None:
        """Probe each mux channel for known device addresses."""
        for mux_idx, tca in enumerate(self.muxes):
            for ch in range(8):
                channel = tca[ch]
                detected: List[str] = []

                if mux_idx == HDC302X_MUX_INDEX and ch == HDC302X_CHANNEL:
                    if self._probe_address(channel, HDC302X_ADDR):
                        detected.append("hdc302x")
                else:
                    if self._probe_address(channel, NAU7802_ADDR):
                        detected.append("nau7802")
                    if self._probe_address(channel, PN532_ADDR):
                        detected.append("pn532")

                if detected:
                    self.active_channels[(mux_idx, ch)] = detected
                    log.info("  mux[%d] ch%d: %s", mux_idx, ch, detected)

    def _disable_unused_channels(self) -> None:
        """Deselect mux channels that have no active devices.
        Lock and immediately unlock each unused channel so that the
        TCA9548A control register no longer routes traffic to it."""
        for mux_idx, tca in enumerate(self.muxes):
            for ch in range(8):
                if (mux_idx, ch) not in self.active_channels:
                    try:
                        ch_obj = tca[ch]
                        while not ch_obj.try_lock():
                            pass
                        ch_obj.unlock()
                    except Exception:
                        pass

    # -- Environmental sensor --

    def _init_env_sensor(self) -> None:
        """Initialize the HDC302x sensor on the designated mux channel."""
        key = (HDC302X_MUX_INDEX, HDC302X_CHANNEL)
        if key not in self.active_channels:
            return
        if HDC302X_MUX_INDEX >= len(self.muxes):
            return
        try:
            tca_channel = self.muxes[HDC302X_MUX_INDEX][HDC302X_CHANNEL]
            self.sensor = adafruit_hdc302x.HDC302x(tca_channel, address=HDC302X_ADDR)
        except Exception:
            self.sensor = None

    def _read_env_sensor(self) -> None:
        """Read temperature and humidity from the HDC302x sensor."""
        if not self.sensor:
            return
        try:
            self.temperature_c = self.sensor.temperature
            self.humidity = self.sensor.relative_humidity
        except Exception:
            self.temperature_c = None
            self.humidity = None

    # -- Slot management --

    def _build_slots(self) -> None:
        """Create FilamentSlot objects for each channel that has a NAU7802.
        slot_id is computed as mux_idx * 8 + channel so that it maps
        to a stable physical location regardless of which other channels
        are detected.  This keeps calibration config keyed correctly
        across reboots even when some channels fail to probe."""
        self._slot_map: Dict[int, FilamentSlot] = {}
        for mux_idx, tca in enumerate(self.muxes):
            for ch in range(8):
                key = (mux_idx, ch)
                devices = self.active_channels.get(key, [])
                if "nau7802" in devices:
                    slot_id = mux_idx * 8 + ch
                    tca_channel = tca[ch]
                    cfg = self.config.get(str(slot_id), SlotConfig())
                    slot = FilamentSlot(slot_id, tca_channel, cfg)
                    self.slots.append(slot)
                    self._slot_map[slot_id] = slot

    def _load_config(self) -> Dict[str, SlotConfig]:
        try:
            if not os.path.exists(CONFIG_FILE):
                return {}
            raw = json.loads(open(CONFIG_FILE, "r", encoding="utf-8").read())
            out: Dict[str, SlotConfig] = {}
            for k, v in (raw.get("slots") or {}).items():
                out[str(k)] = SlotConfig(
                    calibration_factor=float(v.get("calibration_factor", 1.0) or 1.0),
                    zero_offset_raw=float(v.get("zero_offset_raw", 0.0) or 0.0),
                    last_calibrated=v.get("last_calibrated"),
                )
            return out
        except Exception:
            return {}

    def _save_config(self) -> None:
        try:
            payload = {"slots": {str(s.slot_id): {
                "calibration_factor": float(s.cfg.calibration_factor),
                "zero_offset_raw": float(s.cfg.zero_offset_raw),
                "last_calibrated": s.cfg.last_calibrated,
            } for s in self.slots}}
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2)
        except Exception:
            pass

    def get_slot(self, slot_id: int) -> FilamentSlot:
        return self._slot_map[int(slot_id)]

    def update_readings(self) -> None:
        with self.lock:
            try:
                self._read_env_sensor()

                if not hasattr(self, "_nfc_poll_counter"):
                    self._nfc_poll_counter = 0

                for s in self.slots:
                    try:
                        s.refresh_hardware()

                        if not s.hardware_present:
                            s.uid = None
                            s.is_active = False
                            s.gross_weight = 0.0
                            s.weight = 0.0
                            continue

                        # Always read weight
                        s.read_weight()

                        # Only poll NFC occasionally
                        if self._nfc_poll_counter % 5 == 0:
                            s.read_nfc()

                    except Exception as e:
                        s.last_error = f"slot: {e}"

                self._nfc_poll_counter += 1

                self._save_config()
                self.last_error = None

            except Exception as e:
                self.last_error = str(e)

    def snapshot(self) -> Dict[str, Any]:
        with self.lock:
            slots = [s.get_data() for s in self.slots if s.hardware_present]
            return {
                "slots": slots,
                "active_slots": len(slots),
                "present_ports": len(slots),
                "temperature": self.temperature_c,
                "humidity": self.humidity,
                "error": self.last_error,
            }

# ---------------------------
# Flask App + Background Loop
# ---------------------------

app = Flask(__name__)

# Shared I2C bus (CircuitPython)
try:
    _i2c = busio.I2C(board.SCL, board.SDA)
    log.info("I2C bus initialized on SCL=%s SDA=%s", board.SCL, board.SDA)
except Exception as e:
    log.error("Failed to initialize I2C bus: %s", e)
    log.error("Make sure I2C is enabled: sudo raspi-config nonint do_i2c 0")
    sys.exit(1)

try:
    inventory = InventoryManager(_i2c)
    log.info("InventoryManager ready: %d mux(es), %d slot(s), env_sensor=%s",
             len(inventory.muxes), len(inventory.slots), inventory.sensor is not None)
except Exception as e:
    log.error("InventoryManager init failed: %s", e, exc_info=True)
    sys.exit(1)

_bg_started = False
_bg_lock = threading.Lock()

def start_background_loop() -> None:
    global _bg_started
    with _bg_lock:
        if _bg_started:
            return
        _bg_started = True

    def loop():
        while True:
            try:
                inventory.update_readings()
            except Exception:
                pass
            time.sleep(SCAN_INTERVAL_SEC)

    threading.Thread(target=loop, daemon=True).start()

# Start immediately so it works under `flask run` too.
start_background_loop()

# ---------------------------
# Routes
# ---------------------------

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/inventory")
def api_inventory():
    return jsonify(inventory.snapshot()), 200

@app.route("/api/status")
def api_status():
    snap = inventory.snapshot()
    return jsonify({
        "active_slots": snap.get("active_slots", 0),
        "present_ports": snap.get("present_ports", 0),
        "temperature": snap.get("temperature"),
        "humidity": snap.get("humidity"),
        "error": snap.get("error"),
    }), 200

@app.route("/api/program_nfc/<int:slot_id>", methods=["POST"])
def api_program_nfc(slot_id: int):
    data = request.get_json(silent=True) or {}
    brand = clamp_text(data.get("brand") or "")
    color = clamp_text(data.get("color") or "")
    filament_type = clamp_text(data.get("type") or "")
    tare = safe_float(data.get("tare"))

    with inventory.lock:
        slot = inventory.get_slot(slot_id)
        ok = slot.write_tag(brand, color, filament_type, tare)
        append_history({"ts": now_iso(), "slot": slot_id, "event": "program_nfc", "ok": ok, "uid": slot.uid})
        return jsonify({"success": ok, "slot": slot.get_data()}), (200 if ok else 400)

@app.route("/api/rewrite_nfc/<int:slot_id>", methods=["POST"])
def api_rewrite_nfc(slot_id: int):
    return api_program_nfc(slot_id)

@app.route("/api/replace_spool/<int:slot_id>", methods=["POST"])
def api_replace_spool(slot_id: int):
    data = request.get_json(silent=True) or {}
    full_roll_g = float(data.get("full_roll_g") or NEW_ROLL_DEFAULT_G)
    manual_tare = safe_float(data.get("manual_tare"))

    with inventory.lock:
        slot = inventory.get_slot(slot_id)
        slot.read_weight()
        ok = slot.set_new_roll(full_roll_g=full_roll_g, manual_tare=manual_tare)
        append_history({"ts": now_iso(), "slot": slot_id, "event": "replace_spool", "ok": ok, "uid": slot.uid, "full_roll_g": full_roll_g})
        return jsonify({"success": ok, "slot": slot.get_data()}), (200 if ok else 400)


@app.route("/api/zero/<int:slot_id>", methods=["POST"])
def api_zero(slot_id: int):
    """Capture current raw reading as zero offset (no load on scale)."""
    with inventory.lock:
        slot = inventory.get_slot(slot_id)
        if not slot.has_scale:
            return jsonify({"success": False, "error": "Load cell not detected"}), 400
        raw = None
        last_exc = None
        for _ in range(5):
            try:
                raw = slot.nau7802.read_filtered_raw()
                break
            except Exception as e:
                last_exc = e
                time.sleep(0.02)
        if raw is None:
            return jsonify({"success": False, "error": "Unable to read load cell", "details": str(last_exc) if last_exc else None}), 400
        if raw == 0:
            return jsonify({"success": False, "error": "Unable to read load cell"}), 400
        slot.cfg.zero_offset_raw = float(raw)
        inventory._save_config()
        append_history({"ts": now_iso(), "slot": slot_id, "event": "zero", "raw": raw})
        return jsonify({"success": True, "zero_offset_raw": slot.cfg.zero_offset_raw}), 200

@app.route("/api/calibrate/<int:slot_id>", methods=["POST"])
def api_calibrate(slot_id: int):
    data = request.get_json(silent=True) or {}
    known_weight = safe_float(data.get("known_weight"))
    if not known_weight or known_weight <= 0:
        return jsonify({"success": False, "error": "known_weight must be > 0"}), 400

    with inventory.lock:
        slot = inventory.get_slot(slot_id)
        if not slot.has_scale:
            return jsonify({"success": False, "error": "Load cell not detected"}), 400

        raw = slot.nau7802.read_filtered_raw()
        if raw == 0:
            return jsonify({"success": False, "error": "Unable to read load cell"}), 400

        zero = float(slot.cfg.zero_offset_raw or 0.0)
        if zero == 0.0:
            return jsonify({"success": False, "error": "Zero not set. Remove all weight, wait 2-3s, then click Zero Scale."}), 400

        delta = raw - zero
        # NAU7802 raw counts are unitless; some setups yield smaller deltas depending on gain/sample rate.
        # Require a minimum delta, but keep it forgiving for early bring-up.
        if abs(delta) < 200:
            return jsonify({"success": False, "error": "Reading change too small/unstable", "known_weight": known_weight, "raw": raw, "zero": zero, "delta": delta, "hint": "Make sure the known weight is actually on the platform (not touching frame), wait 3-5s, then try again. If delta stays near 0, the NAU7802 is not seeing the bridge signal (wiring/E+/E-/A+/A- or mechanical mount)."}), 400

        factor = delta / float(known_weight)
        if factor == 0:
            return jsonify({"success": False, "error": "Bad calibration factor"}), 400
        # store signed counts-per-gram so direction is preserved
        factor = float(factor)

        slot.cfg.calibration_factor = float(factor)
        slot.cfg.last_calibrated = now_iso()
        inventory._save_config()

        append_history({"ts": now_iso(), "slot": slot_id, "event": "calibrate", "known_weight": known_weight, "factor": slot.cfg.calibration_factor})
        return jsonify({"success": True, "calibration_factor": slot.cfg.calibration_factor, "slot": slot.get_data()}), 200

@app.route("/api/export_csv")
def api_export_csv():
    snap = inventory.snapshot()
    rows = [["Slot","UID","Brand","Type","Color","Gross_g","Tare_g","Net_g","CalStatus","LastCalibrated"]]
    for s in snap.get("slots", []):
        rows.append([
            int(s["slot_id"]) + 1,
            s.get("uid") or "",
            s.get("brand") or "",
            s.get("type") or "",
            s.get("color") or "",
            s.get("gross_weight") or 0,
            s.get("tare") if s.get("tare") is not None else "",
            s.get("weight") or 0,
            s.get("calibration_status") or "",
            s.get("last_calibrated") or "",
        ])

    def esc(v):
        sv = "" if v is None else str(v)
        if any(c in sv for c in [",","\n","\r",'"']):
            sv = sv.replace('"', '""')
            return f'"{sv}"'
        return sv

    csv_text = "\n".join([",".join(esc(c) for c in r) for r in rows])
    buf = io.BytesIO(csv_text.encode("utf-8"))
    return send_file(buf, mimetype="text/csv", as_attachment=True, download_name="inventory_report.csv")

@app.route("/api/export")
def api_export_excel():
    snap = inventory.snapshot()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Slot","UID","Brand","Type","Color","Gross (g)","Tare (g)","Net (g)","Calibration","Last Calibrated"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    for s in snap.get("slots", []):
        ws.append([
            int(s["slot_id"]) + 1,
            s.get("uid") or "",
            s.get("brand") or "",
            s.get("type") or "",
            s.get("color") or "",
            s.get("gross_weight") or 0,
            s.get("tare") if s.get("tare") is not None else "",
            s.get("weight") or 0,
            s.get("calibration_status") or "",
            s.get("last_calibrated") or "",
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="inventory_report.xlsx")

@app.route("/api/export_history")
def api_export_history():
    if not os.path.exists(HISTORY_FILE):
        buf = io.BytesIO(b"")
        return send_file(buf, mimetype="text/plain", as_attachment=True, download_name="history.jsonl")
    return send_file(HISTORY_FILE, mimetype="text/plain", as_attachment=True, download_name="history.jsonl")

# ---------------------------
# Entrypoint
# ---------------------------

if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "80"))
    app.run(host=host, port=port, debug=False)
